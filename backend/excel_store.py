from __future__ import annotations

import threading
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Optional, Set

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.config import OUTPUT_DIR
from src.utils.lead_id import compute_lead_id


_EXCEL_LOCK = threading.Lock()

LEADS_SHEET_NAME = "Leads"

EDITABLE_COLUMNS = [
    "Outreach_Status",
    "Notes",
    "Owner",
    "Last_Contacted",
    "Follow_Up_Date",
]


def get_existing_lead_ids() -> Set[str]:
    """
    Get all Lead_IDs from existing Excel files in the output directory.
    Used for duplicate prevention across scraping runs.
    """
    OUTPUT_DIR.mkdir(exist_ok=True)
    existing_ids: Set[str] = set()
    
    for xlsx_file in OUTPUT_DIR.glob("*.xlsx"):
        try:
            with _EXCEL_LOCK:
                wb = load_workbook(xlsx_file, data_only=True, read_only=True)
                ws = wb[LEADS_SHEET_NAME] if LEADS_SHEET_NAME in wb.sheetnames else wb.active
                
                # Find Lead_ID column
                headers = [cell.value for cell in ws[1]]
                lead_id_col = None
                for idx, header in enumerate(headers):
                    if header == "Lead_ID":
                        lead_id_col = idx
                        break
                
                if lead_id_col is not None:
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and len(row) > lead_id_col and row[lead_id_col]:
                            existing_ids.add(str(row[lead_id_col]))
                
                wb.close()
        except Exception as e:
            # Skip files that can't be read
            continue
    
    return existing_ids


def get_existing_businesses() -> Set[str]:
    """
    Get normalized business identifiers (name + city + state) from existing files.
    Returns set of "business_name|city|state" strings for duplicate checking.
    """
    OUTPUT_DIR.mkdir(exist_ok=True)
    existing: Set[str] = set()
    
    for xlsx_file in OUTPUT_DIR.glob("*.xlsx"):
        try:
            with _EXCEL_LOCK:
                wb = load_workbook(xlsx_file, data_only=True, read_only=True)
                ws = wb[LEADS_SHEET_NAME] if LEADS_SHEET_NAME in wb.sheetnames else wb.active
                
                # Find relevant columns
                headers = [cell.value for cell in ws[1]]
                name_col = city_col = state_col = None
                for idx, header in enumerate(headers):
                    if header == "Business_Name":
                        name_col = idx
                    elif header == "City":
                        city_col = idx
                    elif header == "State":
                        state_col = idx
                
                if name_col is not None:
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and len(row) > name_col and row[name_col]:
                            name = str(row[name_col]).lower().strip()
                            city = str(row[city_col]).lower().strip() if city_col and len(row) > city_col and row[city_col] else ""
                            state = str(row[state_col]).lower().strip() if state_col and len(row) > state_col and row[state_col] else ""
                            existing.add(f"{name}|{city}|{state}")
                
                wb.close()
        except Exception:
            continue
    
    return existing


def _safe_output_path(filename: str) -> Path:
    # Prevent path traversal
    name = Path(filename).name
    path = OUTPUT_DIR / name
    if path.parent.resolve() != OUTPUT_DIR.resolve():
        raise ValueError("Invalid file path")
    return path


def list_output_excel_files() -> list[dict[str, Any]]:
    OUTPUT_DIR.mkdir(exist_ok=True)
    files = sorted(
        OUTPUT_DIR.glob("*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    out: list[dict[str, Any]] = []
    for p in files:
        stat = p.stat()
        out.append(
            {
                "name": p.name,
                "modified_at": datetime.fromtimestamp(stat.st_mtime),
                "size_bytes": stat.st_size,
            }
        )
    return out


def read_leads_from_excel(filename: str) -> tuple[list[str], list[dict[str, Any]]]:
    path = _safe_output_path(filename)
    if not path.exists():
        raise FileNotFoundError(filename)

    with _EXCEL_LOCK:
        wb = load_workbook(path, data_only=True)
        ws = wb[LEADS_SHEET_NAME] if LEADS_SHEET_NAME in wb.sheetnames else wb.active

        headers = _read_headers(ws)
        rows = _read_rows(ws, headers)

    # Always ensure Lead_ID exists in the API response (even for legacy files)
    if "Lead_ID" not in headers:
        headers = ["Lead_ID", *headers]
        for row in rows:
            row["Lead_ID"] = compute_lead_id(
                business_name=row.get("Business_Name") or row.get("Business") or "",
                address=row.get("Address") or "",
                phone=row.get("Phone") or "",
                website=row.get("Website") or "",
            )

    return headers, rows


def update_lead_in_excel(
    filename: str,
    lead_id: str,
    patch: dict[str, Optional[str]],
) -> None:
    path = _safe_output_path(filename)
    if not path.exists():
        raise FileNotFoundError(filename)

    with _EXCEL_LOCK:
        wb = load_workbook(path)
        ws = wb[LEADS_SHEET_NAME] if LEADS_SHEET_NAME in wb.sheetnames else wb.active

        headers = _read_headers(ws)
        col_index = _header_index(headers)

        # Ensure Lead_ID + editable columns exist in the workbook (treat Excel as DB)
        required = ["Lead_ID", *EDITABLE_COLUMNS]
        updated = False
        for col in required:
            if col not in col_index:
                headers.append(col)
                col_index[col] = len(headers)  # 1-based
                ws.cell(row=1, column=col_index[col], value=col)
                updated = True

        # Backfill Lead_IDs if missing/empty
        biz_name_col = col_index.get("Business_Name") or col_index.get("Business")
        address_col = col_index.get("Address")
        phone_col = col_index.get("Phone")
        website_col = col_index.get("Website")
        lead_id_col = col_index["Lead_ID"]

        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=lead_id_col)
            if cell.value:
                continue

            business_name = (
                ws.cell(row=r, column=biz_name_col).value if biz_name_col else ""
            )
            address = ws.cell(row=r, column=address_col).value if address_col else ""
            phone = ws.cell(row=r, column=phone_col).value if phone_col else ""
            website = ws.cell(row=r, column=website_col).value if website_col else ""

            cell.value = compute_lead_id(
                business_name=str(business_name or ""),
                address=str(address or ""),
                phone=str(phone or ""),
                website=str(website or ""),
            )
            updated = True

        # Find row for Lead_ID
        target_row = None
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(row=r, column=lead_id_col).value or "") == lead_id:
                target_row = r
                break

        if not target_row:
            raise KeyError(f"Lead_ID not found: {lead_id}")

        # Apply patch (only editable columns)
        for key, value in patch.items():
            if key not in EDITABLE_COLUMNS:
                continue
            if value is None:
                continue
            ws.cell(row=target_row, column=col_index[key], value=value)
            updated = True

        if updated:
            wb.save(path)


def _read_headers(ws: Worksheet) -> list[str]:
    headers: list[str] = []
    for cell in ws[1]:
        val = cell.value
        if val is None:
            headers.append("")
        else:
            headers.append(str(val).strip())
    # Trim trailing empty header columns
    while headers and headers[-1] == "":
        headers.pop()
    return headers


def _header_index(headers: Iterable[str]) -> dict[str, int]:
    # Map header -> 1-based column index
    out: dict[str, int] = {}
    for idx, h in enumerate(headers, start=1):
        if h:
            out[h] = idx
    return out


def _read_rows(ws: Worksheet, headers: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    max_col = len(headers)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        values = list(row[:max_col])
        # Skip fully-empty rows
        if all(v is None or str(v).strip() == "" for v in values):
            continue
        record: dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            val = values[i] if i < len(values) else None
            record[h] = _json_safe(val)
        rows.append(record)
    return rows


def _json_safe(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    return value

