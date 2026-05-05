"""
Storage module for Lead Prospector v2.
Handles Excel file operations for both cold outreach and direct lead pipelines.
Replaces backend/excel_store.py with multi-directory support.
"""
import threading
from datetime import datetime
from pathlib import Path
from typing import Any, Optional, Set

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.core.config import OUTPUT_DIR, COLD_OUTPUT_DIR, DIRECT_OUTPUT_DIR
from src.utils.lead_id import compute_lead_id

_EXCEL_LOCK = threading.Lock()
LEADS_SHEET_NAME = "Leads"

EDITABLE_COLUMNS = [
    "Outreach_Status",
    "Notes",
    "Owner",
    "Last_Contacted",
    "Follow_Up_Date",
]


def _get_dir(section: str) -> Path:
    """Get the output directory for a section."""
    if section == "cold":
        return COLD_OUTPUT_DIR
    elif section == "direct":
        return DIRECT_OUTPUT_DIR
    elif section == "legacy":
        return OUTPUT_DIR
    raise ValueError(f"Invalid section: {section}")


def list_files(section: str) -> list[dict[str, Any]]:
    """List Excel files for a section (cold, direct, or legacy)."""
    target_dir = _get_dir(section)
    target_dir.mkdir(exist_ok=True)

    if section == "legacy":
        # Only files in root output/, not subdirectories
        files = [f for f in target_dir.glob("*.xlsx") if f.parent == target_dir]
    else:
        files = list(target_dir.glob("*.xlsx"))

    files = sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)
    return [
        {
            "name": p.name,
            "section": section,
            "modified_at": datetime.fromtimestamp(p.stat().st_mtime),
            "size_bytes": p.stat().st_size,
        }
        for p in files
    ]


def _safe_path(filename: str, section: str) -> Path:
    target_dir = _get_dir(section)
    name = Path(filename).name
    path = target_dir / name
    if path.parent.resolve() != target_dir.resolve():
        raise ValueError("Invalid file path")
    return path


def read_leads(filename: str, section: str) -> tuple[list[str], list[dict[str, Any]]]:
    """Read leads from an Excel file. Returns (headers, rows)."""
    path = _safe_path(filename, section)
    if not path.exists():
        raise FileNotFoundError(filename)

    with _EXCEL_LOCK:
        wb = load_workbook(path, data_only=True)
        ws = wb[LEADS_SHEET_NAME] if LEADS_SHEET_NAME in wb.sheetnames else wb.active
        headers = _read_headers(ws)
        rows = _read_rows(ws, headers)

    # Backfill Lead_ID for legacy cold outreach files
    if "Lead_ID" not in headers and section in ("cold", "legacy"):
        headers = ["Lead_ID", *headers]
        for row in rows:
            row["Lead_ID"] = compute_lead_id(
                business_name=row.get("Business_Name") or row.get("Business") or "",
                address=row.get("Address") or "",
                phone=row.get("Phone") or "",
                website=row.get("Website") or "",
            )

    return headers, rows


def update_lead(
    filename: str, lead_id: str, patch: dict[str, Optional[str]], section: str
) -> None:
    """Update a lead's editable fields in an Excel file."""
    path = _safe_path(filename, section)
    if not path.exists():
        raise FileNotFoundError(filename)

    # Migrate outreach status — accept any current Stage enum value, plus legacy aliases.
    if "Outreach_Status" in patch and patch["Outreach_Status"]:
        from src.core.models import Stage
        status = patch["Outreach_Status"].lower().strip()
        # Legacy aliases the codebase has emitted historically:
        _LEGACY_ALIASES = {"queued": Stage.NEW.value, "converted": Stage.WON.value, "passed": Stage.LOST.value}
        valid_stages = {s.value for s in Stage}
        if status in _LEGACY_ALIASES:
            patch["Outreach_Status"] = _LEGACY_ALIASES[status]
        elif status in valid_stages:
            patch["Outreach_Status"] = status
        else:
            patch["Outreach_Status"] = Stage.NEW.value

    with _EXCEL_LOCK:
        wb = load_workbook(path)
        ws = wb[LEADS_SHEET_NAME] if LEADS_SHEET_NAME in wb.sheetnames else wb.active
        headers = _read_headers(ws)
        col_index = _header_index(headers)

        # Ensure required columns exist
        required = ["Lead_ID", *EDITABLE_COLUMNS]
        updated = False
        for col in required:
            if col not in col_index:
                headers.append(col)
                col_index[col] = len(headers)
                ws.cell(row=1, column=col_index[col], value=col)
                updated = True

        # Find lead_id column and target row
        lead_id_col = col_index.get("Lead_ID")
        if not lead_id_col:
            raise KeyError("Lead_ID column not found")

        target_row = None
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(row=r, column=lead_id_col).value or "") == lead_id:
                target_row = r
                break

        if not target_row:
            raise KeyError(f"Lead_ID not found: {lead_id}")

        for key, value in patch.items():
            if key not in EDITABLE_COLUMNS or value is None:
                continue
            ws.cell(row=target_row, column=col_index[key], value=value)
            updated = True

        if updated:
            wb.save(path)


def get_existing_businesses() -> Set[str]:
    """Get normalized business identifiers from existing cold outreach files. For dedup."""
    existing: Set[str] = set()
    # Scan both legacy and cold directories
    for directory in [OUTPUT_DIR, COLD_OUTPUT_DIR]:
        directory.mkdir(exist_ok=True)
        for xlsx_file in directory.glob("*.xlsx"):
            try:
                with _EXCEL_LOCK:
                    wb = load_workbook(xlsx_file, data_only=True, read_only=True)
                    ws = (
                        wb[LEADS_SHEET_NAME]
                        if LEADS_SHEET_NAME in wb.sheetnames
                        else wb.active
                    )
                    headers = [cell.value for cell in ws[1]]
                    name_col = city_col = state_col = None
                    for idx, h in enumerate(headers):
                        if h == "Business_Name":
                            name_col = idx
                        elif h == "City":
                            city_col = idx
                        elif h == "State":
                            state_col = idx
                    if name_col is not None:
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row and len(row) > name_col and row[name_col]:
                                name = str(row[name_col]).lower().strip()
                                city = (
                                    str(row[city_col]).lower().strip()
                                    if city_col
                                    and len(row) > city_col
                                    and row[city_col]
                                    else ""
                                )
                                state = (
                                    str(row[state_col]).lower().strip()
                                    if state_col
                                    and len(row) > state_col
                                    and row[state_col]
                                    else ""
                                )
                                existing.add(f"{name}|{city}|{state}")
                    wb.close()
            except Exception:
                continue
    return existing


def get_existing_direct_lead_urls() -> Set[str]:
    """Get URLs from existing direct lead files. For dedup."""
    DIRECT_OUTPUT_DIR.mkdir(exist_ok=True)
    urls: Set[str] = set()
    for xlsx_file in DIRECT_OUTPUT_DIR.glob("*.xlsx"):
        try:
            with _EXCEL_LOCK:
                wb = load_workbook(xlsx_file, data_only=True, read_only=True)
                ws = (
                    wb[LEADS_SHEET_NAME]
                    if LEADS_SHEET_NAME in wb.sheetnames
                    else wb.active
                )
                headers = [cell.value for cell in ws[1]]
                url_col = None
                for idx, h in enumerate(headers):
                    if h and h.lower() in ("url", "job_url", "post_url"):
                        url_col = idx
                        break
                if url_col is not None:
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and len(row) > url_col and row[url_col]:
                            urls.add(str(row[url_col]).strip())
                wb.close()
        except Exception:
            continue
    return urls


# Private helpers (same as existing excel_store.py)


def _read_headers(ws: Worksheet) -> list[str]:
    headers = []
    for cell in ws[1]:
        val = cell.value
        headers.append(str(val).strip() if val is not None else "")
    while headers and headers[-1] == "":
        headers.pop()
    return headers


def _header_index(headers) -> dict[str, int]:
    out = {}
    for idx, h in enumerate(headers, start=1):
        if h:
            out[h] = idx
    return out


def _read_rows(ws: Worksheet, headers: list[str]) -> list[dict[str, Any]]:
    rows = []
    max_col = len(headers)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        values = list(row[:max_col])
        if all(v is None or str(v).strip() == "" for v in values):
            continue
        record = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            val = values[i] if i < len(values) else None
            if isinstance(val, datetime):
                record[h] = val.isoformat()
            else:
                record[h] = val
        rows.append(record)
    return rows
